from tkinter import *

import yaml
from openpyxl import load_workbook

from gui import GUI, DetailsWindow
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
from datetime import datetime
from os.path import basename

##########################
#       EMAIL DATA       #
##########################
class GeneratedEmail:
    static_id = 0
    def __init__(self, to, cc, subject, attachments, body_raw, row):
        self.to = to
        self.cc = cc
        self.subject = subject
        self.body_raw = body_raw
        self.body_generated = body_raw
        self.row = row
        self.attachments = attachments
        GeneratedEmail.static_id += 1
        self.id = GeneratedEmail.static_id
    def parse_body(self, excel_data, data_config):
        self.body_generated = self.body_raw

        #sorting data to stop subset conflicts
        initial_dictionary = {}
        for i in data_config:
            initial_dictionary.update(i)
        data_values = {}
        temp = list(initial_dictionary.items())
        temp.sort(key=lambda x:len(x[0]),reverse=True)
        for i in temp:
            data_values.update({i[0]:i[1]})
        
        #replacing values
        for data in data_values:
            data_converted = str(excel_data.cell(row=self.row, column=data_values[data]).value)
            self.body_generated = self.body_generated.replace("%"+data, data_converted)
    def generate_mimetext(self, sender):
        #Email formatting
        mimetext_formatted = MIMEMultipart()
        mimetext_formatted.attach(MIMEText(self.body_generated, 'html'))
        mimetext_formatted["From"] = sender
        mimetext_formatted["Subject"] = self.subject
        mimetext_formatted["To"] = ", ".join(self.to)
        mimetext_formatted["CC"] = ", ".join(self.cc)
        for f in files or []:
            with open(f, "rb") as fil:
                part = MIMEApplication(
                    fil.read(),
                    Name=basename(f)
                )
            # After the file is closed
            part['Content-Disposition'] = 'attachment; filename="%s"' % basename(f)
            mimetext_formatted.attach(part)
        return mimetext_formatted
    def __str__(self):
        return "to:  {0}, cc: {1}, subject: {2}".format(self.to, self.cc, self.subject)


class Controller:
    def __init__(self, gui):
        self.initialise()
        #gui
        self.gui = gui
        self.gui.controller = self
        self.logging = Logging()
        

    def initialise(self):
        #server
        self.smtp_server = None
    
        #authentication status
        self.authentication_status = ""
        #Data configs
        self.email_config = None

        #
        self.workbook = None
        self.worksheet = None 

        #email data
        self.generated_email_list = []

        self.logging = Logging()
        

    def load_yaml(self, filename):
        if(filename != None and filename != ""):
            self.initialise()
            self.gui.reset()
            with open(filename,'r') as file:
                self.email_config = yaml.load(file, Loader=yaml.FullLoader)
            self.gui.html_editor.delete('1.0', END) 
            self.gui.html_editor.insert(END,self.email_config["EMAIL_TEMPLATE"])
            self.load_excel_data()
            self.generated_email()
            self.gui.master.title(self.gui.default_title + " - "+str(len(self.generated_email_list)) + " emails loaded")
    def init_sending_emails_authentication(self):
        if(len(self.generated_email_list) > 0):
            self.gui.display_details_window()
    
    def calculate_normalised_range(self, range_config):
        return range(int(range_config["START"]), int(range_config["END"])+1)

    def check_if_null(self, value):
        return value == None or value == "None" or value == ""

    def bulk_send_emails(self, username, password):
        if(self.init_server()):
            if(self.login(username,password)):
                for email in self.generated_email_list:
                    email.body_raw = self.gui.html_editor.get("1.0",'end-1c')
                    email.parse_body(self.worksheet, self.email_config["DATA"])
                    self.send_email(email, username)
                #email completion window
                self.gui.display_job_completion_widow()
                self.gui.job_completion_window.update_job_completion_status()
                #can request report
                self.gui.mainmenu.entryconfig(3, state="normal" )

    def send_email(self, email, username):
        formatted_email = email.generate_mimetext(username)
        try:
            self.smtp_server.sendmail(username, formatted_email["To"]+","+formatted_email["CC"],formatted_email.as_string())
            if(self.gui.progress_bar['value'] + 100.0/len(self.generated_email_list) >= 100):
                self.gui.progress_bar['value'] = 100.0
            else:
                self.gui.update_progress_bar(100.0/len(self.generated_email_list))
            self.logging.emails_success.append(email)
            return True
        except:
            self.logging.emails_failed.append(email)
            return False

    def init_server(self):
        self.gui.details_window.login_button.config(state='disabled')
        server = self.email_config["SERVER"]
        port = int(self.email_config["PORT"])
        try:
            self.gui.details_window.update_authentication_status_label("Connecting to " + server + " on port: "+str(port))
            self.smtp_server = smtplib.SMTP(server, port)
            self.smtp_server.connect(server, port)
            self.smtp_server.ehlo()
            self.smtp_server.starttls()
            self.smtp_server.ehlo()
            return True
        except Exception as e:
            print(e)
            self.gui.details_window.update_authentication_status_label("Connection failed")
            self.gui.details_window.login_button.config(state='normal')
            return False

    def login(self, username, password):
        try:
            self.gui.details_window.update_authentication_status_label("Connection Success. Attempting login")
            self.smtp_server.login(username, password)
            self.gui.details_window.window.destroy()
            return True
        except Exception as e:
            print(e)
            self.gui.details_window.update_authentication_status_label("Login failed")
            self.reset_server()
            self.gui.details_window.login_button.config(state='normal')
            return False
    
    def reset_server(self):
        self.smtp_server.close()

    def get_recipients_from_excel(self, row, recipient):
        if(not self.check_if_null(self.email_config[recipient[1]])):
            for recipient_type_list in self.email_config[recipient[1]]:
                if(not self.check_if_null(list(recipient_type_list.values())[0])):
                    value_in_cell = str(self.worksheet.cell(row=row, column=list(recipient_type_list.values())[0]).value)
                    if(not self.check_if_null(value_in_cell)):
                        recipient[0].append(value_in_cell)
                        
    def generated_email(self):
        #email generation
        self.generated_email_list = []
        for row in self.calculate_normalised_range(self.email_config["RANGE"]):
            #recipients details
            to = []
            cc = []
            recipient_list = [(to, "TO"), (cc, "CC")]
            for recipient in recipient_list:
                self.get_recipients_from_excel(row, recipient)

            #attachment details
            attachments = []
            attachment_list = [(attachments, "ATTACHMENT")]
            for attachment in attachment_list:
                self.get_recipients_from_excel(row, attachment)
                
            #email details
            subject = self.email_config["SUBJECT"]
            body_template = self.email_config["EMAIL_TEMPLATE"]
            generated_email = GeneratedEmail(to,cc,subject,attachments,body_template, row)
            self.generated_email_list.append(generated_email)
        self.gui.update_list(self.generated_email_list)

    def load_excel_data(self):
        self.workbook = load_workbook(self.email_config["FILE_DIRECTORY"])
        self.worksheet = self.workbook[self.email_config["SHEET_NAME"]]
    def initialise_GUI(self):
        self.gui.initialise()

class Logging:
    def __init__(self):
        self.emails_failed = []
        self.emails_success = []
    def output_to_file(self):
        try:
            file = open("email_report_"+str(datetime.now().strftime("%Y_%m_%d__%H_%M_%S")+".txt"),'x')
            file.write("emails failed"+"\n")
            for email in self.emails_failed:
                file.write(str(email)+"\n")
            file.write("\n\n\n\n")
            file.write("emails success"+"\n")
            for email in self.emails_success:
                file.write(str(email)+"\n")
            file.close()
        except:
            pass

def main():
    #GUI    
    root = Tk()
    root.geometry("960x720+150+150")
    gui = GUI(root)
    controller = Controller(gui)
    controller.initialise_GUI()
    root.mainloop()

if __name__ == "__main__":
    main()
